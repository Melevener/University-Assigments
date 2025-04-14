import openpyxl

#Открытие таблицы
wb = openpyxl.load_workbook("35вариант.xlsx")

#Работа с листом "Исполнители"
executor = wb["Исполнители"]

data_executor = []
executor_cod = 0
data_executor_kod = []

# Сортировка листа "Исполнитель": Нужен код исполнителя старше 30 лет с Китая
for row in executor.iter_rows(min_row=2, values_only=True):
    data_executor.append(row)
    for tuple_row in data_executor:
        if tuple_row[2] == "Китай" and tuple_row[1] > 30:
            executor_cod = tuple_row[0]
            data_executor_kod.append(executor_cod)
unique_data_executor_kod = set(data_executor_kod)

#Работа с листом "Услуги"
service = wb["Услуги"]

service_cod = 0

for row in service.iter_rows(min_row=2, values_only=True):
    if row[1] == "Python-программист":
        service_cod = row[0]

#Работа с листом "Заказы"
order = wb["Заказы"]

order_cost = 0
data_order_cost = []
for row in order.iter_rows(min_row=2, values_only=True):
    if row[1] == 11 and row[2] in unique_data_executor_kod:
        order_cost = row[3]
        data_order_cost.append(order_cost)

for row in order.iter_rows(min_row=2, values_only=True):
    if row[3] == max(data_order_cost):
        print(row[2])




