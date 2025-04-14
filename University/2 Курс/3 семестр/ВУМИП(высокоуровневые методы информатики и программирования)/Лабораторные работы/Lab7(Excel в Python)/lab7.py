import openpyxl

#Инициализация переменных
Condition1 = False
Condition2 = False
Condition3 = False
count_solve_string = 0

#Открытие таблицы
wb = openpyxl.load_workbook("35вариант.xlsx")
sheet = wb.active

#Нахождение количества строк
count_row = sheet.max_row

#Проверка условий для строки
for row in sheet.iter_rows(min_row=1, max_row=count_row, values_only=True):
    #Условие 1: только одно число встречается в строке дважды
    unique_numbers = set(row)
    if len(unique_numbers) == 3:
        Condition1 = True

    #Условие 2: сумма двух самых больших чисел строки более чем в два раза больше суммы двух самых малых
    sorted_row = sorted(row, reverse=True)
    if sorted_row[0] + sorted_row[1] > 2 * (sorted_row[2] + sorted_row[3]):
        Condition2 = True

    #Условие 3: максимальное число строки не кратно минимальному
    if sorted_row[0] % sorted_row[3] != 0:
        Condition3 = True

    #Проверка всех 3-х условий и добавление в счетчик
    if Condition1 and Condition2 and Condition3:
        count_solve_string += 1

    Condition1 = False
    Condition2 = False
    Condition3 = False
print(count_solve_string)
